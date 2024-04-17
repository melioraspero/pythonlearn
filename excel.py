import pandas as pd

# Шаг 1: Чтение данных из файла Excel
df = pd.read_excel('имя_файла.xlsx')

# Шаг 2: Обработка данных (необязательно)
# ...

# Шаг 3: Запись данных в файл Excel
df.to_excel('новый_файл.xlsx', index=False)  # Установите index=False, чтобы не сохранять индексы DataFrame
В этом примере сначала мы читаем данные из файла Excel в DataFrame с помощью функции read_excel(), затем обрабатываем данные (если это необходимо), и наконец записываем измененные данные в новый файл Excel с помощью метода to_excel().

пример для получения записей из книги, листа и таблицы Excel, а затем выгрузим их на простую таблицу веб-страницы.

python
Copy code
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, render_template

# Создаем объект приложения Flask
app = Flask(__name__)

# Функция для чтения данных из книги Excel
def read_excel_data(file_path, sheet_name):
    # Загружаем книгу Excel
    workbook = load_workbook(file_path)
    # Получаем лист по имени
    sheet = workbook[sheet_name]
    # Получаем данные из таблицы (предполагаем, что они начинаются с первой строки и первого столбца)
    data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=1, max_row=200, max_col=3)]
    # Возвращаем данные в формате DataFrame
    return pd.DataFrame(data, columns=['Column1', 'Column2', 'Column3'])

# Определяем маршрут для отображения данных на веб-странице
@app.route('/')
def display_data():
    # Путь к файлу Excel и имя листа
    file_path = 'имя_файла.xlsx'
    sheet_name = 'Лист1'
    # Читаем данные из книги Excel
    df = read_excel_data(file_path, sheet_name)
    # Преобразуем DataFrame в HTML-таблицу
    table_html = df.to_html(index=False)
    # Отображаем HTML-таблицу на веб-странице
    return render_template('index.html', table=table_html)

# Запускаем сервер Flask
if __name__ == '__main__':
    app.run(debug=True)
В этом примере мы используем Flask для создания веб-приложения. Функция read_excel_data() читает данные из книги Excel и возвращает их в виде DataFrame. Затем мы используем метод to_html() DataFrame, чтобы преобразовать данные в HTML-таблицу, которую мы отображаем на веб-странице с помощью Jinja2.







  
Библиотека openpyxl:
python
Copy code
from openpyxl import load_workbook

# Шаг 1: Загрузка книги Excel
workbook = load_workbook('имя_файла.xlsx')

# Шаг 2: Получение активного листа
sheet = workbook.active

# Шаг 3: Чтение/изменение данных (необязательно)
# ...

# Шаг 4: Сохранение книги
workbook.save('новый_файл.xlsx')
Здесь мы открываем книгу Excel с помощью load_workbook(), получаем активный лист и вносим необходимые изменения в данные (если это необходимо), а затем сохраняем книгу с помощью метода save().







Библиотека xlrd/xlwt:
python
Copy code
import xlrd
import xlwt

# Шаг 1: Чтение данных из файла Excel
workbook = xlrd.open_workbook('имя_файла.xlsx')
sheet = workbook.sheet_by_index(0)

# Шаг 2: Получение данных (необязательно)
# ...

# Шаг 3: Создание новой книги Excel и листа
new_workbook = xlwt.Workbook()
new_sheet = new_workbook.add_sheet('Новый лист')

# Шаг 4: Запись данных в новую книгу Excel (необязательно)
# ...

# Шаг 5: Сохранение новой книги Excel
new_workbook.save('новый_файл.xlsx')
Здесь мы сначала открываем файл Excel с помощью xlrd.open_workbook(), читаем данные (если это необходимо), создаем новую книгу Excel с помощью xlwt.Workbook() и добавляем в нее новый лист. Затем мы записываем данные в новую книгу (если это необходимо) и сохраняем ее с помощью метода save().
